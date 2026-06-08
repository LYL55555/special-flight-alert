"""Special-livery tables: cumulative history archive + horizon snapshot for new/expired."""

from __future__ import annotations

import shutil
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from alerts.run_paths import AlertRunPaths
from alerts.scorer import AlertExtras
from alerts.snapshot_report import (
    SNAPSHOT_XLSX_COLUMNS,
    diff_qualifying,
    qualifying_rows,
    read_snapshot_keys,
    read_snapshot_rows_by_key,
    write_qualifying_xlsx,
)
from config import EngineConfig
from models.flight import Flight

LIVERY_ARCHIVE_EXTRA_COLUMNS: Tuple[str, ...] = (
    "first_seen_run",
    "last_seen_run",
    "times_seen",
    "archive_key",
)

LIVERY_ARCHIVE_COLUMNS: Tuple[str, ...] = tuple(
    c for c in SNAPSHOT_XLSX_COLUMNS if c != "snapshot_key"
) + LIVERY_ARCHIVE_EXTRA_COLUMNS


def is_special_livery(extras: AlertExtras, reasons: List[str]) -> bool:
    if (extras.livery_name or "").strip():
        return True
    return any(r.startswith("Special livery:") for r in reasons)


def filter_livery_qualifying(
    qualifying: List[Tuple[Flight, int, List[str], AlertExtras]],
) -> List[Tuple[Flight, int, List[str], AlertExtras]]:
    return [
        item
        for item in qualifying
        if is_special_livery(item[3], item[2])
    ]


def archive_key(row: Dict[str, Any]) -> str:
    reg = (row.get("registration") or "").strip().upper()
    paint = (row.get("livery_name") or "").strip()
    if reg and paint:
        return f"reg:{reg}|paint:{paint}"
    return f"leg:{row.get('snapshot_key') or ''}"


def _archive_row_from_snapshot(row: Dict[str, Any], *, run_ts: str) -> Dict[str, Any]:
    out = {k: row.get(k) for k in SNAPSHOT_XLSX_COLUMNS if k != "snapshot_key"}
    key = archive_key(row)
    out["archive_key"] = key
    out["first_seen_run"] = run_ts
    out["last_seen_run"] = run_ts
    out["times_seen"] = 1
    return out


def _read_archive_rows(path: Path) -> Dict[str, Dict[str, Any]]:
    from openpyxl import load_workbook

    if not path.is_file():
        return {}
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if not header:
            return {}
        names = [str(h) if h is not None else "" for h in header]
        out: Dict[str, Dict[str, Any]] = {}
        for row in it:
            if not row:
                continue
            d = {names[i]: row[i] for i in range(min(len(names), len(row)))}
            k = str(d.get("archive_key") or "").strip()
            if not k:
                k = archive_key(d)
            if k and k != "leg:":
                out[k] = d
        return out
    finally:
        wb.close()


def merge_livery_archive(
    existing: Dict[str, Dict[str, Any]],
    horizon_rows: List[Dict[str, Any]],
    *,
    run_ts: str,
) -> List[Dict[str, Any]]:
    merged = dict(existing)
    for row in horizon_rows:
        key = archive_key(row)
        if not key or key == "leg:":
            continue
        if key in merged:
            prev = merged[key]
            times = int(prev.get("times_seen") or 1) + 1
            updated = {k: row.get(k) for k in SNAPSHOT_XLSX_COLUMNS if k != "snapshot_key"}
            updated["archive_key"] = key
            updated["first_seen_run"] = prev.get("first_seen_run") or run_ts
            updated["last_seen_run"] = run_ts
            updated["times_seen"] = times
            merged[key] = updated
        else:
            merged[key] = _archive_row_from_snapshot(row, run_ts=run_ts)
    return sorted(
        merged.values(),
        key=lambda r: (
            str(r.get("last_seen_run") or ""),
            str(r.get("spot_time_local") or ""),
        ),
        reverse=True,
    )


def _write_archive_xlsx(path: Path, rows: List[Dict[str, Any]]) -> None:
    normalized = [{c: r.get(c) for c in LIVERY_ARCHIVE_COLUMNS} for r in rows]
    write_qualifying_xlsx(path, normalized, columns=LIVERY_ARCHIVE_COLUMNS)


def update_livery_archive_by_airport(
    qualifying: List[Tuple[Flight, int, List[str], AlertExtras]],
    config: EngineConfig,
    run_paths: AlertRunPaths,
) -> List[Path]:
    """Append/update cumulative special-livery sightings per airport (reference log)."""
    livery = filter_livery_qualifying(qualifying)
    groups: Dict[str, List[Tuple[Flight, int, List[str], AlertExtras]]] = defaultdict(list)
    for item in livery:
        ap = (item[0].monitored_airport or "UNK").strip().upper() or "UNK"
        groups[ap].append(item)

    written: List[Path] = []
    for ap in sorted(groups.keys()):
        rows = qualifying_rows(groups[ap], config, run_paths=run_paths)
        history_path = run_paths.livery_history_path(ap)
        existing = _read_archive_rows(history_path)
        merged = merge_livery_archive(existing, rows, run_ts=run_paths.run_ts)
        _write_archive_xlsx(history_path, merged)
        written.append(history_path)
    return written


def update_livery_horizon_by_airport(
    qualifying: List[Tuple[Flight, int, List[str], AlertExtras]],
    config: EngineConfig,
    run_paths: AlertRunPaths,
    *,
    max_lines_per_airport: Optional[int] = None,
) -> Tuple[int, int, List[str], List[str], int, List[Path]]:
    """
    Horizon special-livery snapshot per airport; diff vs _latest drives new/expired.
    """
    livery = filter_livery_qualifying(qualifying)
    groups: Dict[str, List[Tuple[Flight, int, List[str], AlertExtras]]] = defaultdict(list)
    for item in livery:
        ap = (item[0].monitored_airport or "UNK").strip().upper() or "UNK"
        groups[ap].append(item)

    tot_exp = 0
    tot_new = 0
    all_exp: List[str] = []
    all_new: List[str] = []
    written: List[Path] = []

    for ap in sorted(groups.keys()):
        chunk = groups[ap]
        rows = qualifying_rows(chunk, config, run_paths=run_paths)
        latest = run_paths.livery_snapshot_latest_path(ap)
        stamp = run_paths.livery_snapshot_run_path(ap)
        old_keys = read_snapshot_keys(latest)
        old_by_k = read_snapshot_rows_by_key(latest)
        n_e, n_n, el, nl = diff_qualifying(
            old_keys,
            rows,
            old_rows_by_key=old_by_k,
            max_lines=max_lines_per_airport,
        )
        tot_exp += n_e
        tot_new += n_n
        all_exp.extend(el)
        all_new.extend(nl)
        write_qualifying_xlsx(stamp, rows)
        shutil.copyfile(stamp, latest)
        written.append(stamp)

    return tot_exp, tot_new, all_exp, all_new, len(livery), written


def update_livery_horizon_single(
    config: EngineConfig,
    qualifying: List[Tuple[Flight, int, List[str], AlertExtras]],
) -> Tuple[int, int, List[str], List[str], int]:
    """Legacy single-file layout (--output mode)."""
    data_dir = Path(config.schedule_snapshot_xlsx_path).parent
    latest = data_dir / "livery_snapshot_latest.xlsx"
    livery = filter_livery_qualifying(qualifying)
    rows = qualifying_rows(livery, config)
    old_keys = read_snapshot_keys(latest)
    old_by_k = read_snapshot_rows_by_key(latest)
    n_exp, n_new, expired_lines, new_lines = diff_qualifying(
        old_keys, rows, old_rows_by_key=old_by_k
    )
    write_qualifying_xlsx(latest, rows)
    return n_exp, n_new, expired_lines, new_lines, len(livery)


def update_livery_archive_single(
    config: EngineConfig,
    qualifying: List[Tuple[Flight, int, List[str], AlertExtras]],
    *,
    run_ts: str,
) -> Path:
    data_dir = Path(config.schedule_snapshot_xlsx_path).parent
    history = data_dir / "livery_history.xlsx"
    livery = filter_livery_qualifying(qualifying)
    rows = qualifying_rows(livery, config)
    existing = _read_archive_rows(history)
    merged = merge_livery_archive(existing, rows, run_ts=run_ts)
    _write_archive_xlsx(history, merged)
    return history
