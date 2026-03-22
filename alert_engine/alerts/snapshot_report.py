"""Excel snapshot of qualifying flights + diff vs previous run."""

from __future__ import annotations

import re
import shutil
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from alerts.run_paths import AlertRunPaths
from alerts.scorer import AlertExtras
from config import EngineConfig
from models.flight import Flight

JETPHOTOS_REGISTRATION_URL = "https://www.jetphotos.com/registration/{}"


def jetphotos_url_for_registration(reg: Optional[str]) -> str:
    """https://www.jetphotos.com/registration/N586JB or .../HB-IFA"""
    r = (reg or "").strip()
    if not r or r in ("?", "N/A", "n/a"):
        return ""
    r = r.upper().replace(" ", "")
    if len(r) < 2:
        return ""
    return JETPHOTOS_REGISTRATION_URL.format(r)


# Excel export: spot time first, snapshot id last; no FR24 / schedule id / paths.
SNAPSHOT_XLSX_COLUMNS: Tuple[str, ...] = (
    "spot_time_local",
    "monitored_airport",
    "movement",
    "flight_number",
    "registration",
    "jetphotos_url",
    "aircraft_type",
    "operator",
    "airline_icao",
    "origin",
    "destination",
    "scheduled_departure_local",
    "estimated_departure_local",
    "scheduled_arrival_local",
    "estimated_arrival_local",
    "departure_tz",
    "arrival_tz",
    "livery_name",
    "livery_airline",
    "livery_description",
    "score",
    "reasons",
    "source",
    "snapshot_key",
)


def snapshot_key(flight: Flight) -> str:
    rid = (flight.schedule_row_id or "").strip()
    if rid:
        return f"sched:{rid}"
    return f"leg:{flight.leg_signature()}"


def _strip_operator_paren(op: str) -> str:
    s = (op or "").strip()
    if not s:
        return ""
    t = re.sub(r"\s*\([^)]*\)", "", s).strip()
    return t or s


def _row_dict(
    flight: Flight,
    score: int,
    reasons: List[str],
    extras: AlertExtras,
    config: EngineConfig,
    *,
    run_paths: Optional[AlertRunPaths] = None,
) -> Dict[str, Any]:
    _ = (config, run_paths)  # callers may pass; snapshot rows omit paths / ids
    tloc = flight.row_local_times()
    key = snapshot_key(flight)
    return {
        "spot_time_local": flight.spot_time_local_display(),
        "monitored_airport": flight.monitored_airport or "",
        "movement": flight.movement or "",
        "flight_number": flight.flight_number or "",
        "registration": flight.registration or "",
        "jetphotos_url": jetphotos_url_for_registration(flight.registration),
        "aircraft_type": flight.aircraft_type or "",
        "operator": flight.operator or "",
        "airline_icao": flight.airline_icao or "",
        "origin": flight.origin or "",
        "destination": flight.destination or "",
        "scheduled_departure_local": tloc["scheduled_departure_local"],
        "estimated_departure_local": tloc["estimated_departure_local"],
        "scheduled_arrival_local": tloc["scheduled_arrival_local"],
        "estimated_arrival_local": tloc["estimated_arrival_local"],
        "departure_tz": tloc["departure_tz_label"],
        "arrival_tz": tloc["arrival_tz_label"],
        "livery_name": extras.livery_name,
        "livery_airline": extras.livery_airline,
        "livery_description": extras.livery_description,
        "score": score,
        "reasons": " | ".join(reasons),
        "source": flight.source,
        "snapshot_key": key,
    }


def format_digest_line(row: Dict[str, Any]) -> str:
    """One line for Telegram (list item body without leading '- ')."""
    spot = (row.get("spot_time_local") or "—") if row else "—"
    op_full = (row.get("operator") or row.get("airline_icao") or "").strip() or "—"
    short = (row.get("livery_airline") or "").strip() or _strip_operator_paren(op_full)
    if not short or short == "—":
        short = op_full
    fn = row.get("flight_number") or "?"
    reg = row.get("registration") or "?"
    ap = row.get("monitored_airport") or "?"
    mv = row.get("movement") or "?"
    base = f"[{spot}] {short} {fn} {reg} ｜ {ap} {mv} | {op_full}"
    jp = (row.get("jetphotos_url") or "").strip() or jetphotos_url_for_registration(
        str(row.get("registration") or "")
    )
    if jp:
        return f"{base} | {jp}"
    return base


def row_dict_for_export(row: Dict[str, Any]) -> Dict[str, Any]:
    return {k: row.get(k) for k in SNAPSHOT_XLSX_COLUMNS}


def qualifying_rows(
    qualifying: List[Tuple[Flight, int, List[str], AlertExtras]],
    config: EngineConfig,
    *,
    run_paths: Optional[AlertRunPaths] = None,
) -> List[Dict[str, Any]]:
    return [
        row_dict_for_export(_row_dict(f, s, r, e, config, run_paths=run_paths))
        for f, s, r, e in qualifying
    ]


def _read_xlsx_rows(xlsx_path: Path) -> Tuple[List[str], List[Dict[str, Any]]]:
    if not xlsx_path.is_file():
        return [], []
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if not header:
            return [], []
        names = [str(h) if h is not None else "" for h in header]
        rows_out: List[Dict[str, Any]] = []
        for row in it:
            if not row:
                continue
            d = {names[i]: row[i] for i in range(min(len(names), len(row)))}
            if d.get("snapshot_key"):
                rows_out.append(d)
        return names, rows_out
    finally:
        wb.close()


def read_snapshot_keys(xlsx_path: Path) -> Set[str]:
    _, rows = _read_xlsx_rows(xlsx_path)
    return {str(r["snapshot_key"]).strip() for r in rows if r.get("snapshot_key")}


def read_snapshot_rows_by_key(xlsx_path: Path) -> Dict[str, Dict[str, Any]]:
    _, rows = _read_xlsx_rows(xlsx_path)
    out: Dict[str, Dict[str, Any]] = {}
    for r in rows:
        k = str(r.get("snapshot_key") or "").strip()
        if k:
            out[k] = r
    return out


def diff_qualifying(
    old_keys: Set[str],
    new_rows: List[Dict[str, Any]],
    *,
    old_rows_by_key: Dict[str, Dict[str, Any]] | None = None,
    max_lines: Optional[int] = None,
) -> Tuple[int, int, List[str], List[str]]:
    new_map = {r["snapshot_key"]: r for r in new_rows}
    new_keys = set(new_map)
    expired_keys = old_keys - new_keys
    fresh_keys = new_keys - old_keys
    n_exp, n_new = len(expired_keys), len(fresh_keys)
    prev = old_rows_by_key or {}
    expired_lines: List[str] = []
    for k in sorted(expired_keys):
        r = prev.get(k)
        expired_lines.append(format_digest_line(r) if r else k)
    new_lines = [format_digest_line(new_map[k]) for k in sorted(fresh_keys)]
    if max_lines is not None:
        if len(expired_lines) > max_lines:
            extra = len(expired_lines) - max_lines
            expired_lines = expired_lines[:max_lines] + [f"... +{extra} more"]
        if len(new_lines) > max_lines:
            extra = len(new_lines) - max_lines
            new_lines = new_lines[:max_lines] + [f"... +{extra} more"]
    return n_exp, n_new, expired_lines, new_lines


def write_qualifying_xlsx(path: Path, rows: List[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "qualifying"
    headers = list(SNAPSHOT_XLSX_COLUMNS)
    if not rows:
        ws.append(headers)
        _autofit_columns(ws, headers, [])
        wb.save(path)
        return
    ws.append(headers)
    for ri, r in enumerate(rows, start=2):
        for ci, h in enumerate(headers, start=1):
            val = r.get(h)
            c = ws.cell(row=ri, column=ci)
            if h == "jetphotos_url" and val:
                s = str(val)
                c.value = s
                if s.startswith("http"):
                    c.hyperlink = s
                    c.font = Font(color="0563C1", underline="single")
            else:
                c.value = val
    _autofit_columns(ws, headers, rows)
    wb.save(path)


def _autofit_columns(
    ws: Any,
    headers: List[str],
    rows: List[Dict[str, Any]],
    *,
    min_w: float = 14.0,
    max_w: float = 56.0,
) -> None:
    for i, h in enumerate(headers, start=1):
        letter = get_column_letter(i)
        cap = 72.0 if h == "jetphotos_url" else max_w
        maxlen = len(str(h))
        for r in rows:
            v = r.get(h)
            if v is None:
                continue
            maxlen = max(maxlen, len(str(v)))
        ws.column_dimensions[letter].width = max(min_w, min(cap, maxlen + 2.5))


def update_snapshots_by_airport(
    qualifying: List[Tuple[Flight, int, List[str], AlertExtras]],
    config: EngineConfig,
    run_paths: AlertRunPaths,
    *,
    max_lines_per_airport: Optional[int] = None,
) -> Tuple[int, int, List[str], List[str], int, List[Path]]:
    """
    Per monitored_airport: diff vs snapshot_{AP}_latest.xlsx, write snapshot_{AP}_{run_ts}.xlsx,
    then copy to _latest for the next run.
    """
    groups: Dict[str, List[Tuple[Flight, int, List[str], AlertExtras]]] = defaultdict(list)
    for item in qualifying:
        ap = (item[0].monitored_airport or "UNK").strip().upper() or "UNK"
        groups[ap].append(item)

    tot_exp = 0
    tot_new = 0
    all_exp: List[str] = []
    all_new: List[str] = []
    written: List[Path] = []

    for ap in sorted(groups.keys()):
        chunk = groups[ap]
        rows = qualifying_rows(chunk, config, run_paths=run_paths)
        latest = run_paths.snapshot_latest_path(ap)
        stamp = run_paths.snapshot_run_path(ap)
        old_keys = read_snapshot_keys(latest)
        old_by_k = read_snapshot_rows_by_key(latest)
        n_e, n_n, el, nl = diff_qualifying(
            old_keys,
            rows,
            old_rows_by_key=old_by_k,
            max_lines=max_lines_per_airport,
        )
        tot_exp += n_e
        tot_new += n_n
        all_exp.extend(el)
        all_new.extend(nl)
        write_qualifying_xlsx(stamp, rows)
        shutil.copyfile(stamp, latest)
        written.append(stamp)

    return tot_exp, tot_new, all_exp, all_new, len(qualifying), written


def update_single_snapshot(
    config: EngineConfig,
    qualifying: List[Tuple[Flight, int, List[str], AlertExtras]],
) -> Tuple[int, int, List[str], List[str], int]:
    """Legacy: one global xlsx path from config."""
    snap_path = Path(config.schedule_snapshot_xlsx_path)
    old_keys = read_snapshot_keys(snap_path)
    old_by_k = read_snapshot_rows_by_key(snap_path)
    rows = qualifying_rows(qualifying, config)
    n_exp, n_new, expired_lines, new_lines = diff_qualifying(
        old_keys,
        rows,
        old_rows_by_key=old_by_k,
        max_lines=None,
    )
    write_qualifying_xlsx(snap_path, rows)
    return n_exp, n_new, expired_lines, new_lines, len(qualifying)


def sort_row_dicts_for_display(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    def k(r: Dict[str, Any]) -> Tuple[str, str, str]:
        return (
            str(r.get("monitored_airport") or ""),
            str(r.get("spot_time_local") or ""),
            str(r.get("flight_number") or ""),
        )

    return sorted(rows, key=k)
